from openpyxl import Workbook
# wb = Workbook()

# grab the active worksheet
# ws = wb.active

# Data can be assigned directly to cells
# ws['A1'] = 42

# Rows can also be appended
# ws.append([1, 2, 3])

# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# Save the file
# wb.save("sample.xlsx")


# from openpyxl import load_workbook
# wb = load_workbook(filename = "sample.xlsx")
# sheet_ranges = wb['Sheet']
# print(sheet_ranges['A1'].value)


# import openpyxl
# wb = openpyxl.Workbook()
# ws = wb.create_sheet()
# ws.column_dimensions.group('A','D', hidden=True)
# ws.row_dimensions.group(1,10, hidden=True)
# wb.save('group.xlsx')


from openpyxl import Workbook
from openpyxl.utils import get_column_letter
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)